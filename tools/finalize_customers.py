#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tools/finalize_customers.py — gộp kết quả cuối sau khi người dùng điền "Chủ chọn".

    python3 tools/finalize_customers.py <Customers_Merged_đã_điền.xlsx> \
        <Customer.xlsx> <Customer_Data.xlsx> <out.xlsx>

Đọc lựa chọn tay ở sheet "Cần rà" (cột "Chủ chọn" dạng "Tên pháp nhân · Chủ"),
áp lên list, rồi TỰ LỌC TRÙNG: pháp nhân nào đã được nhận là một khách trong list
thì bỏ khỏi "Khách mới".

Xuất:
  Cập nhật    — 293 khách của list, Owner/LegalName điền tới đâu biết tới đó.
  Khách mới   — khách trong file CHƯA có trong list (đã bỏ trùng), đúng cột list.
  Còn trống   — khách trong list vẫn chưa có chủ (điền dần sau).
"""
import sys, re, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill

PREFIXES = [
    "VĂN PHÒNG ĐẠI DIỆN", "VPĐD",
    "CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI DỊCH VỤ", "CÔNG TY TNHH THƯƠNG MẠI DỊCH VỤ",
    "CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI", "CÔNG TY TNHH THƯƠNG MẠI",
    "CÔNG TY TNHH SẢN XUẤT", "CÔNG TY TNHH MỘT THÀNH VIÊN", "CÔNG TY TNHH MTV",
    "CÔNG TY CỔ PHẦN SẢN XUẤT THƯƠNG MẠI", "CÔNG TY CỔ PHẦN THƯƠNG MẠI",
    "CÔNG TY CỔ PHẦN TẬP ĐOÀN", "CÔNG TY CỔ PHẦN", "CÔNG TY TNHH", "CÔNG TY CP",
    "CÔNG TY", "CTY TNHH", "CTY CP", "CTY", "TỔNG CÔNG TY",
    "DOANH NGHIỆP TƯ NHÂN", "DNTN", "HỘ KINH DOANH", "HKD", "CHI NHÁNH", "NHÀ MÁY",
    "CƠ SỞ SẢN XUẤT", "CƠ SỞ", "XÍ NGHIỆP", "HỢP TÁC XÃ", "HTX",
    "TNHH MTV", "TNHH", "MTV", "CP",
]
BOUND = " \t-–—:.,;()"
STOP = {"CONG","TY","TNHH","CO","CP","PHAN","SAN","XUAT","THUONG","MAI","DICH","VU",
        "VIET","NAM","GROUP","FOOD","FOODS","JSC","LTD","MTV","HO","KINH","DOANH",
        "CHI","NHANH","NHA","MAY","SO","VAN","PHONG","DAI","DIEN","TAP","DOAN","AND"}


def dz(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D").upper()


PN = [dz(p) for p in PREFIXES]


def strip_one(name):
    t = name.lstrip(BOUND); head = dz(t)
    for p in PN:
        if head.startswith(p):
            a = head[len(p):len(p)+1]
            if a == "" or a in BOUND:
                return t[len(p):]
    return None


def clean_name(raw):
    s = re.sub(r"\s+", " ", str(raw or "")).strip()
    if not s: return ""
    cur = s
    for _ in range(6):
        st = strip_one(cur)
        if st is None: break
        st = re.sub(r"\s+", " ", st.lstrip(BOUND)).strip()
        if not st: break
        cur = st
    return cur or s


def okey(raw):
    return re.sub(r"\s+", " ", dz(clean_name(raw))).strip()


def norm_full(s): return re.sub(r"[^A-Z0-9]+", " ", dz(s)).strip()
def toks(s): return [t for t in norm_full(s).split() if t and t not in STOP and len(t) > 1]
def phrase(hay, needle):
    return bool(needle) and re.search(r"(?:^| )"+re.escape(needle)+r"(?:$| )", hay) is not None


def col_index(header, *names):
    low = {str(c.value).strip().lower(): i+1 for i, c in enumerate(header) if c.value}
    for n in names:
        if n.strip().lower() in low: return low[n.strip().lower()]
    return None


def main():
    merged, export, owners, out = sys.argv[1], sys.argv[2], sys.argv[3], \
        (sys.argv[4] if len(sys.argv) > 4 else "Customers_Final.xlsx")

    # --- file chủ sở hữu ---
    ob = openpyxl.load_workbook(owners, read_only=True, data_only=True)
    ows = ob[ob.sheetnames[0]]
    file_rows, owner_of, legal_of = [], {}, {}
    for r in ows.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]: continue
        legal = str(r[0]).strip(); owner = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        k = okey(legal)
        if k and k not in owner_of: owner_of[k] = owner; legal_of[k] = legal
        file_rows.append((legal, owner, norm_full(legal), k))

    def auto_match(title):
        k = okey(title)
        if k in owner_of: return owner_of[k], legal_of[k], k
        nt = norm_full(title)
        hits = [(lg, ow, kk) for lg, ow, nf, kk in file_rows if phrase(nf, nt)]
        if hits:
            os = set(o for _, o, _ in hits if o)
            if len(os) == 1: return next(iter(os)), hits[0][0], hits[0][2]
        return "", "", ""

    # --- lựa chọn tay từ sheet "Cần rà" ---
    mb = openpyxl.load_workbook(merged, data_only=True)
    choice_owner, choice_legal = {}, {}     # theo okey(title)
    if "Cần rà" in mb.sheetnames:
        cr = mb["Cần rà"]
        for r in cr.iter_rows(min_row=2, values_only=True):
            title = r[0]; pick = r[4] if len(r) > 4 else None
            if not title or not (pick and str(pick).strip()): continue
            parts = [p.strip() for p in str(pick).split("·")]
            if len(parts) >= 2:
                legal, owner = parts[0], parts[-1]
            else:
                legal, owner = "", parts[0]
            choice_owner[okey(title)] = owner
            if legal: choice_legal[okey(title)] = legal

    # --- list export ---
    eb = openpyxl.load_workbook(export, data_only=True)
    es = eb[eb.sheetnames[0]]
    header = list(es[1])
    ci_t = col_index(header, "Title", "Tên khách hàng")
    SYS = {"item type", "path", "id", "content type", "modified", "created"}
    cols = [(i+1, str(c.value)) for i, c in enumerate(header)
            if c.value and str(c.value).strip().lower() not in SYS]
    names = [n for _, n in cols]
    if "Owner" not in names: names.append("Owner")
    if "LegalName" not in names: names.append("LegalName")

    used = set()                # okey pháp nhân đã gắn vào một khách của list
    out_wb = openpyxl.Workbook()
    up = out_wb.active; up.title = "Cập nhật"; up.append(names)
    for c in up[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="01426A")
    green = PatternFill("solid", fgColor="D1FAE5")
    yellow = PatternFill("solid", fgColor="FEF9C3")

    blanks = []
    for row in es.iter_rows(min_row=2, values_only=False):
        title = row[ci_t-1].value
        if not title: continue
        base = {n: (es.cell(row=row[0].row, column=ci).value if ci else "")
                for ci, n in cols}
        kt = okey(title)
        owner = legal = ""; how = ""
        if kt in choice_owner:                       # ưu tiên lựa chọn tay
            owner = choice_owner[kt]; legal = choice_legal.get(kt, ""); how = "manual"
            if legal: used.add(okey(legal))
        else:
            owner, legal, mk = auto_match(title)
            if owner:
                how = "auto"; used.add(mk)
        rowvals = []
        for n in names:
            low = n.strip().lower()
            if low in ("owner", "người phụ trách"): rowvals.append(owner)
            elif low in ("legalname", "tên pháp nhân"): rowvals.append(legal or base.get(n, ""))
            else: rowvals.append(base.get(n, ""))
        up.append(rowvals)
        if how:
            fill = green if how == "manual" else yellow
            for n in ("Owner", "LegalName"):
                if n in names:
                    up.cell(row=up.max_row, column=names.index(n)+1).fill = fill
        else:
            blanks.append(title)

    # --- Khách mới: file chưa dùng, đúng cột list ---
    nw = out_wb.create_sheet("Khách mới"); nw.append(names)
    for c in nw[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="157F3C")
    n_new = 0
    for k, owner in owner_of.items():
        if k in used: continue
        row = []
        for n in names:
            low = n.strip().lower()
            if low == "title": row.append(clean_name(legal_of[k]))
            elif low in ("owner", "người phụ trách"): row.append(owner)
            elif low in ("legalname", "tên pháp nhân"): row.append(legal_of[k])
            else: row.append("")
        nw.append(row); n_new += 1

    # --- Còn trống ---
    bl = out_wb.create_sheet("Còn trống"); bl.append(["Title (chưa có chủ)"])
    bl["A1"].font = Font(bold=True)
    for t in blanks: bl.append([t])

    for sh in (up, nw):
        for i, n in enumerate(names, 1):
            sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
                40 if n.strip().lower() in ("title", "legalname", "tên pháp nhân") else 18
        sh.freeze_panes = "A2"

    out_wb.save(out)
    have = 293 - len(blanks)
    print("Đã xuất:", out)
    print("  List: 293 khách · đã có chủ:", have, "· còn trống:", len(blanks))
    print("  Lựa chọn tay áp vào:", len(choice_owner))
    print("  Khách mới (đã bỏ trùng với list):", n_new)
    print("  Trùng đã lọc khỏi Khách mới:", len([k for k in choice_legal.values()]))


if __name__ == "__main__":
    main()
