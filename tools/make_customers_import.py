#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tools/make_customers_import.py — sinh file import list Customers từ file thô.

    python3 tools/make_customers_import.py <Customer_Data.xlsx> <out.xlsx>

Thuật toán làm gọn tên PHẢI khớp js/lib/cleanname.py-tương-đương (cleanCustomerName).
Xuất 4 cột:
  Tên gọn (Title) · Tên pháp nhân (LegalName) · Chủ sở hữu (Owner) · Tên gốc
"Tên gốc" chỉ để bạn đối chiếu — bỏ cột đó trước khi import cũng được.

Không ghi gì lên SharePoint. Việc import do bạn làm tay, sau khi duyệt.
"""
import sys, re, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ---- tiền tố pháp nhân, DÀI trước NGẮN (khớp js/lib/cleanname.js) ----
PREFIXES = [
    "VĂN PHÒNG ĐẠI DIỆN", "VPĐD",
    "CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI DỊCH VỤ",
    "CÔNG TY TNHH THƯƠNG MẠI DỊCH VỤ",
    "CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI",
    "CÔNG TY TNHH THƯƠNG MẠI",
    "CÔNG TY TNHH SẢN XUẤT",
    "CÔNG TY TNHH MỘT THÀNH VIÊN",
    "CÔNG TY TNHH MTV",
    "CÔNG TY CỔ PHẦN SẢN XUẤT THƯƠNG MẠI",
    "CÔNG TY CỔ PHẦN THƯƠNG MẠI",
    "CÔNG TY CỔ PHẦN TẬP ĐOÀN",
    "CÔNG TY CỔ PHẦN",
    "CÔNG TY TNHH",
    "CÔNG TY CP",
    "CÔNG TY",
    "CTY TNHH", "CTY CP", "CTY",
    "TỔNG CÔNG TY",
    "DOANH NGHIỆP TƯ NHÂN", "DNTN",
    "HỘ KINH DOANH", "HKD",
    "CHI NHÁNH",
    "NHÀ MÁY",
    "CƠ SỞ SẢN XUẤT", "CƠ SỞ",
    "XÍ NGHIỆP",
    "HỢP TÁC XÃ", "HTX",
    "TNHH MTV", "TNHH", "MTV", "CP",
]
BOUND = " \t-–—:.,;()"


def strip_diacritics(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D")


PREFIX_NORM = [(p, strip_diacritics(p).upper()) for p in PREFIXES]


def strip_one(name):
    trimmed = name.lstrip(BOUND)
    head = strip_diacritics(trimmed).upper()
    for _raw, norm in PREFIX_NORM:
        if head.startswith(norm):
            after = head[len(norm): len(norm) + 1]
            if after == "" or after in BOUND:
                return trimmed[len(norm):]
    return None


def clean_name(raw):
    s = re.sub(r"\s+", " ", str(raw or "")).strip()
    if not s:
        return ""
    cur = s
    for _ in range(6):
        step = strip_one(cur)
        if step is None:
            break
        step = re.sub(r"\s+", " ", step.lstrip(BOUND)).strip()
        if not step:
            break
        cur = step
    return cur or s


def owner_key(raw):
    return re.sub(r"\s+", " ", strip_diacritics(clean_name(raw)).upper()).strip()


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "Customer_Data.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "Customers_Import.xlsx"
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [(str(r[0]).strip(), str(r[1]).strip() if r[1] else "")
            for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

    seen, cross = {}, {}
    for legal, owner in rows:
        k = owner_key(legal)
        seen.setdefault(k, set()).add(owner)
    for k, owners in seen.items():
        owners.discard("")
        if len(owners) > 1:
            cross[k] = owners

    ob = openpyxl.Workbook()
    sh = ob.active
    sh.title = "Customers"
    head = ["Tên gọn (Title)", "Tên pháp nhân (LegalName)", "Chủ sở hữu (Owner)", "Tên gốc (đối chiếu)"]
    sh.append(head)
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="01426A")
        c.alignment = Alignment(vertical="center")
    warn = PatternFill("solid", fgColor="FDE68A")
    for legal, owner in rows:
        title = clean_name(legal)
        sh.append([title, legal, owner, legal])
        if owner_key(legal) in cross:               # cùng tên gọn, khác chủ → tô vàng
            for c in sh[sh.max_row]:
                c.fill = warn
    widths = [40, 52, 22, 52]
    for i, wdt in enumerate(widths, 1):
        sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
    sh.freeze_panes = "A2"
    ob.save(out)

    print("Đã xuất:", out)
    print("  Tổng dòng:", len(rows))
    print("  Tên gọn riêng:", len(seen))
    print("  Nhóm cùng tên gọn nhưng khác chủ (tô vàng):", len(cross))
    for k, owners in list(cross.items())[:20]:
        print("    ", k, "→", ", ".join(sorted(owners)))


if __name__ == "__main__":
    main()
