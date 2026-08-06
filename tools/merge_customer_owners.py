#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""tools/merge_customer_owners.py — ghép chủ sở hữu vào list Customers ĐANG CÓ,
GIỮ NGUYÊN mọi cột khác (Segment, Region, CustomerStatus…).

    python3 tools/merge_customer_owners.py <export_Customers.xlsx> <Customer_Data.xlsx> <out.xlsx>

  · export_Customers.xlsx : bản XUẤT list Customers hiện tại từ SharePoint
                            (Export to Excel). Phải có cột Title.
  · Customer_Data.xlsx    : file chủ sở hữu gốc (Tên khách hàng, Chủ sở hữu).
  · out.xlsx              : file kết quả để bạn cập nhật lại.

Ghép theo custOwnerKey (tên đã bỏ tiền tố pháp nhân + bỏ dấu). Kết quả:
  Sheet "Cập nhật"  — Y NGUYÊN thứ tự & cột của bản xuất, chỉ ĐIỀN thêm Owner và
                      LegalName vào ô đang trống (không đè giá trị bạn đã có).
                      Dán lại cột Owner (và LegalName) qua Edit in grid view.
  Sheet "Khách mới"  — khách trong file chủ sở hữu mà list hiện tại CHƯA có.
  Sheet "Chưa khớp"  — khách trong list hiện tại mà file không có chủ → để trống,
                      giữ phân quyền theo PIC dự án.

KHÔNG đụng SharePoint. KHÔNG đè dữ liệu bạn đã nhập tay. Chỉ điền chỗ trống.
"""
import sys, re, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ---- thuật toán làm gọn (khớp js/lib/cleanname.js & make_customers_import.py) ----
PREFIXES = [
    "VĂN PHÒNG ĐẠI DIỆN", "VPĐD",
    "CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI DỊCH VỤ", "CÔNG TY TNHH THƯƠNG MẠI DỊCH VỤ",
    "CÔNG TY TNHH SẢN XUẤT THƯƠNG MẠI", "CÔNG TY TNHH THƯƠNG MẠI",
    "CÔNG TY TNHH SẢN XUẤT", "CÔNG TY TNHH MỘT THÀNH VIÊN", "CÔNG TY TNHH MTV",
    "CÔNG TY CỔ PHẦN SẢN XUẤT THƯƠNG MẠI", "CÔNG TY CỔ PHẦN THƯƠNG MẠI",
    "CÔNG TY CỔ PHẦN TẬP ĐOÀN", "CÔNG TY CỔ PHẦN", "CÔNG TY TNHH", "CÔNG TY CP",
    "CÔNG TY", "CTY TNHH", "CTY CP", "CTY", "TỔNG CÔNG TY",
    "DOANH NGHIỆP TƯ NHÂN", "DNTN", "HỘ KINH DOANH", "HKD", "CHI NHÁNH", "NHÀ MÁY",
    "CƠ SỞ SẢN XUẤT", "CƠ SỞ", "XÍ NGHIỆP", "HỢP TÁC XÃ", "HTX",
    "TNHH MTV", "TNHH", "MTV", "CP",
]
BOUND = " \t-–—:.,;()"


def strip_diacritics(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D")


PREFIX_NORM = [strip_diacritics(p).upper() for p in PREFIXES]


def strip_one(name):
    t = name.lstrip(BOUND)
    head = strip_diacritics(t).upper()
    for norm in PREFIX_NORM:
        if head.startswith(norm):
            after = head[len(norm):len(norm) + 1]
            if after == "" or after in BOUND:
                return t[len(norm):]
    return None


def clean_name(raw):
    s = re.sub(r"\s+", " ", str(raw or "")).strip()
    if not s:
        return ""
    cur = s
    for _ in range(6):
        step = strip_one(cur)
        if step is None:
            break
        step = re.sub(r"\s+", " ", step.lstrip(BOUND)).strip()
        if not step:
            break
        cur = step
    return cur or s


def okey(raw):
    return re.sub(r"\s+", " ", strip_diacritics(clean_name(raw)).upper()).strip()


# ---- khớp mờ giữa tên ngắn (list) và tên pháp nhân đầy đủ (file) ----
STOP = {"CONG", "TY", "TNHH", "CO", "CP", "PHAN", "SAN", "XUAT", "THUONG", "MAI",
        "DICH", "VU", "VIET", "NAM", "GROUP", "FOOD", "FOODS", "JSC", "LTD",
        "MTV", "HO", "KINH", "DOANH", "CHI", "NHANH", "NHA", "MAY", "CO", "SO",
        "VAN", "PHONG", "DAI", "DIEN", "TAP", "DOAN", "AND"}


def norm_full(s):
    return re.sub(r"[^A-Z0-9]+", " ", strip_diacritics(s).upper()).strip()


def tokens(s):
    return [t for t in norm_full(s).split() if t and t not in STOP and len(t) > 1]


def contains_phrase(hay, needle):
    """needle (đã chuẩn hoá) xuất hiện như cụm từ có ranh giới trong hay."""
    if not needle:
        return False
    return re.search(r"(?:^| )" + re.escape(needle) + r"(?:$| )", hay) is not None


def col_index(header_row, *names):
    """Tìm cột theo tên (không phân biệt hoa thường / khoảng trắng). Trả 1-based."""
    low = {str(c.value).strip().lower(): i + 1 for i, c in enumerate(header_row) if c.value}
    for n in names:
        if n.strip().lower() in low:
            return low[n.strip().lower()]
    return None


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    export = sys.argv[1]
    owners = sys.argv[2]
    out = sys.argv[3] if len(sys.argv) > 3 else "Customers_Merged.xlsx"

    # --- bảng chủ sở hữu từ file gốc ---
    ob = openpyxl.load_workbook(owners, read_only=True, data_only=True)
    ows = ob[ob.sheetnames[0]]
    owner_of, legal_of = {}, {}
    file_rows = []                       # [(legal, owner, norm_full, key)]
    for r in ows.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        legal = str(r[0]).strip()
        owner = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        k = okey(legal)
        if k and k not in owner_of:
            owner_of[k] = owner
            legal_of[k] = legal
        file_rows.append((legal, owner, norm_full(legal), k))

    def match_owner(title):
        """Ghép một tên trong list với chủ sở hữu trong file.
        Trả về (owner, legal, cách_khớp, ứng_viên_gợi_ý).
        cách_khớp: 'exact' | 'substring' | '' (không chắc)."""
        k = okey(title)
        if k in owner_of:
            return owner_of[k], legal_of[k], "exact", []
        nt = norm_full(title)
        # tên list xuất hiện như cụm từ trong tên pháp nhân đầy đủ
        hits = [(lg, ow) for lg, ow, nf, _ in file_rows if contains_phrase(nf, nt)]
        if hits:
            owners = set(o for _, o in hits if o)
            if len(owners) == 1:                    # mọi dòng khớp cùng một chủ
                ow = next(iter(owners))
                return ow, hits[0][0], "substring", []
            # nhiều chủ khác nhau → mơ hồ, đưa ra gợi ý
            return "", "", "", [lg + " · " + ow for lg, ow in hits[:5]]
        # gợi ý để RÀ TAY, không tự gán. Chỉ đưa gợi ý ĐÁNG TIN, nếu không thà để
        # trống còn hơn gợi ý sai (vd "Huu Binh" vs "Vinasoy Bình Dương" chỉ trùng
        # "BINH"). Tiêu chí: trùng ≥2 token đặc trưng, HOẶC trùng đúng 1 token dài
        # (≥4 ký tự) mà token đó hiếm trong file (≤3 dòng).
        tset = set(tokens(title))
        if not tset:
            return "", "", "", []
        # tần suất token trong file để loại token phổ biến
        freq = {}
        for _, _, nf, _ in file_rows:
            for t in set(tokens(nf)):
                freq[t] = freq.get(t, 0) + 1
        scored = []
        for lg, ow, nf, _ in file_rows:
            inter = tset & set(tokens(lg))
            if not inter:
                continue
            strong = [t for t in inter if len(t) >= 4]
            rare = [t for t in strong if freq.get(t, 0) <= 3]
            if len(inter) >= 2 or (len(strong) >= 1 and rare):
                score = len(inter) * 10 + sum(len(t) for t in strong)
                scored.append((score, lg, ow))
        scored.sort(reverse=True)
        return "", "", "", [lg + " · " + ow for _, lg, ow in scored[:3]]

    # --- bản xuất Customers hiện tại ---
    eb = openpyxl.load_workbook(export, data_only=True)
    es = eb[eb.sheetnames[0]]
    header = list(es[1])
    ci_title = col_index(header, "Title", "Tên", "Tên khách hàng")
    if not ci_title:
        print("Không tìm thấy cột Title trong bản xuất. Cột hiện có:",
              [c.value for c in header])
        sys.exit(2)
    ci_owner = col_index(header, "Owner", "Người phụ trách")
    ci_legal = col_index(header, "LegalName", "Tên pháp nhân")

    # nếu bản xuất chưa có cột Owner/LegalName thì thêm vào cuối
    if not ci_owner:
        ci_owner = es.max_column + 1
        es.cell(row=1, column=ci_owner, value="Owner")
    if not ci_legal:
        ci_legal = es.max_column + 1
        es.cell(row=1, column=ci_legal, value="LegalName")

    fill_exact = PatternFill("solid", fgColor="D1FAE5")   # xanh: khớp chắc
    fill_sub = PatternFill("solid", fgColor="FEF9C3")     # vàng nhạt: khớp mờ, nên rà
    matched, submatch, kept = 0, 0, 0
    review = []      # [(title, gợi ý…)] — không gán được chắc
    used = set()
    for row in range(2, es.max_row + 1):
        title = es.cell(row=row, column=ci_title).value
        if not title:
            continue
        owner, legal, how, cands = match_owner(title)
        if how:                                          # gán được
            used.add(okey(legal) if legal else "")
            cur_owner = es.cell(row=row, column=ci_owner).value
            fill = fill_exact if how == "exact" else fill_sub
            if not (cur_owner and str(cur_owner).strip()):
                es.cell(row=row, column=ci_owner, value=owner).fill = fill
                if how == "exact":
                    matched += 1
                else:
                    submatch += 1
            else:
                kept += 1
            cur_legal = es.cell(row=row, column=ci_legal).value
            if not (cur_legal and str(cur_legal).strip()) and legal:
                es.cell(row=row, column=ci_legal, value=legal).fill = fill
        else:
            review.append((str(title), cands))

    es.title = "Cập nhật"

    # sheet khách mới (có trong file chủ, chưa có trong list)
    new_sheet = eb.create_sheet("Khách mới")
    new_sheet.append(["Tên gọn (Title)", "Tên pháp nhân (LegalName)", "Chủ sở hữu (Owner)"])
    for c in new_sheet[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="01426A")
    n_new = 0
    for k, owner in owner_of.items():
        if k in used:
            continue
        new_sheet.append([clean_name(legal_of[k]), legal_of[k], owner])
        n_new += 1

    # sheet cần rà: list-customer chưa gán được chắc, kèm gợi ý từ file
    rv_sheet = eb.create_sheet("Cần rà")
    rv_sheet.append(["Title (trong list)", "Gợi ý 1", "Gợi ý 2", "Gợi ý 3",
                     "Chủ chọn (điền tay)"])
    for c in rv_sheet[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B45309")
    for title, cands in review:
        rv_sheet.append([title] + (cands + ["", "", ""])[:3] + [""])
    for i, w in enumerate([26, 40, 40, 40, 22], 1):
        rv_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    rv_sheet.freeze_panes = "A2"

    eb.save(out)
    print("Đã xuất:", out)
    print("  Khớp CHẮC (tên trùng sau khi bỏ tiền tố):", matched)
    print("  Khớp MỜ  (tên list nằm trong tên pháp nhân, 1 chủ):", submatch)
    print("  Đã có chủ, giữ nguyên:", kept)
    print("  CẦN RÀ TAY (kèm gợi ý):", len(review))
    print("  Khách mới (có trong file, chưa có trong list):", n_new)
    print("\nSheet 'Cập nhật': ô XANH = khớp chắc, ô VÀNG = khớp mờ nên liếc lại.")
    print("Sheet 'Cần rà': mỗi dòng kèm tối đa 3 gợi ý; điền cột 'Chủ chọn' rồi dán về.")
    print("Sheet 'Khách mới': thêm nếu muốn danh bạ đủ khách từ file.")


if __name__ == "__main__":
    main()
